# flask_dashboard.py - Complete Fixed Version
import os
import io
import time
import threading
import datetime as dt
import json
import smtplib
import logging
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
from functools import wraps

from flask import (
    Flask, request, redirect, url_for, render_template, send_file,
    jsonify, flash, Response, session, make_response, send_from_directory, abort
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from sqlalchemy import and_, text
from werkzeug.security import generate_password_hash, check_password_hash

try:
    from dotenv import load_dotenv, find_dotenv
    load_dotenv(find_dotenv(), override=True)
except Exception:
    pass

from db import SessionLocal, User, Attendance, init_db

# ============================================================================
# CONFIGURATION
# ============================================================================

class Config:
    APP_TITLE = os.getenv("APP_TITLE", "Smart Attendance System")
    SECRET_KEY = os.getenv("FLASK_SECRET", os.urandom(24).hex())
    ENV = os.getenv("FLASK_ENV", "production")
    DEBUG = ENV == "development"
    
    SESSION_COOKIE_SECURE = ENV == "production"
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = "Lax"
    PERMANENT_SESSION_LIFETIME = 3600
    
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    UPLOAD_DIR = os.path.abspath(os.getenv("UPLOAD_DIR", os.path.join(BASE_DIR, "uploads")))
    SEED_DIR = os.path.abspath(os.getenv("SEED_DIR", os.path.join(BASE_DIR, "seed_images")))
    STATIC_DIR = os.path.join(BASE_DIR, "static")
    TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
    
    ADMIN_PASSWORD_HASH = generate_password_hash(os.getenv("ADMIN_PASSWORD", "admin123"))
    
    ENGINE_FRAME_PATH = os.getenv("ENGINE_FRAME_PATH", os.path.join(STATIC_DIR, "last_frame.jpg"))
    ENGINE_STATUS_PATH = os.getenv("ENGINE_STATUS_PATH", os.path.join(STATIC_DIR, "last_status.json"))
    ENABLE_CAMERA_STREAM = bool(int(os.getenv("ENABLE_CAMERA_STREAM", "1")))
    
    SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
    SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    SMTP_USER = os.getenv("SMTP_USER")
    SMTP_PASS = os.getenv("SMTP_PASS")
    SMTP_FROM = os.getenv("SMTP_FROM", "attendance@company.com")
    ENABLE_EMAIL = bool(SMTP_USER and SMTP_PASS)
    
    RATELIMIT_STORAGE_URL = os.getenv("REDIS_URL", "memory://")
    RATELIMIT_ENABLED = ENV == "production"

for d in [Config.UPLOAD_DIR, Config.SEED_DIR, Config.STATIC_DIR]:
    os.makedirs(d, exist_ok=True)

# ============================================================================
# LOGGING
# ============================================================================

logging.basicConfig(
    level=logging.INFO if Config.ENV == "production" else logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.FileHandler('dashboard.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# FLASK APP
# ============================================================================

app = Flask(__name__, template_folder=Config.TEMPLATE_DIR, static_folder=Config.STATIC_DIR)
app.config.from_object(Config)

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per hour"],
    storage_uri=Config.RATELIMIT_STORAGE_URL,
    enabled=Config.RATELIMIT_ENABLED
)

# ============================================================================
# CONSTANTS
# ============================================================================

GATES: Dict[int, Tuple[str, str, str]] = {
    1: ("Check-in 1 (Morning Start)", "07:00", "09:00"),
    2: ("Check-in 2 (Lunch Break Out)", "12:00", "13:30"),
    3: ("Check-in 3 (Lunch Break In)", "13:00", "14:30"),
    4: ("Check-in 4 (Evening End)", "17:00", "19:30"),
}

LATE_AT: Dict[int, str] = {1: "08:15", 2: "12:40", 3: "13:10", 4: "17:10"}

# ============================================================================
# DECORATORS
# ============================================================================

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            flash("Please log in to access this page", "warning")
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Admin access required", "error")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return decorated

# ============================================================================
# SECURITY HEADERS
# ============================================================================

@app.after_request
def set_security_headers(response):
    if Config.ENV == "production":
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'SAMEORIGIN'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

# ============================================================================
# HELPERS
# ============================================================================

def today_str() -> str:
    return dt.date.today().isoformat()

def parse_hhmm(s: str) -> dt.time:
    h, m = map(int, s.split(":")[:2])
    return dt.time(h, m)

def is_late(gate_id: int, tstr: Optional[str]) -> bool:
    if not tstr:
        return False
    late_cut = parse_hhmm(LATE_AT.get(gate_id, "23:59"))
    t = dt.datetime.strptime(tstr, "%H:%M:%S").time() if ":" in tstr else parse_hhmm(tstr)
    return t > late_cut

def quarter_for(d: dt.date) -> str:
    q = (d.month - 1) // 3 + 1
    return f"{d.year}{q}"

def send_email(to_addr: Optional[str], subject: str, body_html: str):
    if not Config.ENABLE_EMAIL or not to_addr:
        logger.debug(f"Email disabled or no recipient: {to_addr}")
        return
    
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = f"{Config.APP_TITLE} <{Config.SMTP_FROM}>"
        msg["To"] = to_addr
        html_part = MIMEText(body_html, "html", "utf-8")
        msg.attach(html_part)
        
        with smtplib.SMTP(Config.SMTP_HOST, Config.SMTP_PORT, timeout=10) as s:
            s.starttls()
            s.login(Config.SMTP_USER, Config.SMTP_PASS)
            s.sendmail(Config.SMTP_FROM, [to_addr], msg.as_string())
        
        logger.info(f"Email sent to {to_addr}: {subject}")
    except Exception as e:
        logger.error(f"Email error to {to_addr}: {e}")

def send_welcome_email(name: str, email: str):
    subject = f"Welcome to {Config.APP_TITLE}"
    body = f"""
    <html><body style="font-family:Arial;padding:20px;background:#f5f5f5;">
    <div style="max-width:600px;margin:0 auto;background:#fff;padding:24px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
        <h2 style="color:#3b82f6;margin:0 0 16px;border-bottom:2px solid #3b82f6;padding-bottom:8px;">Welcome to {Config.APP_TITLE}</h2>
        <p>Hi <strong>{name}</strong>,</p>
        <p>Your account has been successfully created in our attendance system.</p>
        <p><strong>What's Next:</strong></p>
        <ul>
            <li>Your face has been enrolled for automatic recognition</li>
            <li>Check in at designated times daily:
                <ul>
                    <li>Check-in 1: Morning Start (07:00-09:00)</li>
                    <li>Check-in 2: Lunch Break Out (12:00-13:30)</li>
                    <li>Check-in 3: Lunch Break In (13:00-14:30)</li>
                    <li>Check-in 4: Evening End (17:00-19:30)</li>
                </ul>
            </li>
            <li>Stand in front of the camera when checking in</li>
            <li>You'll receive notifications for late arrivals</li>
        </ul>
        <p>If you have any questions, please contact your administrator.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="color:#666;font-size:12px;">Automated message • {Config.APP_TITLE}</p>
    </div></body></html>
    """
    send_email(email, subject, body)

def email_template(kind: str, user_name: str, when: str = None):
    if kind == "late":
        subject = "⏰ Late Arrival Notice"
        body = f"""
        <html><body style="font-family:Arial;padding:20px;background:#f5f5f5;">
        <div style="max-width:600px;margin:0 auto;background:#fff;padding:24px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
            <h2 style="color:#ff9800;margin:0 0 16px;border-bottom:2px solid #ff9800;padding-bottom:8px;">Late Arrival Notification</h2>
            <p>Hi <strong>{user_name}</strong>,</p>
            <p>Your morning check-in was recorded at <strong style="color:#ff5722;">{when}</strong>,
            which is after the scheduled time of <strong>{LATE_AT[1]}</strong>.</p>
            <p>Please ensure timely arrival to maintain productivity.</p>
            <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
            <p style="color:#666;font-size:12px;">Automated message • {Config.APP_TITLE}</p>
        </div></body></html>
        """
    else:
        subject = "⚠️ Absence Notice"
        body = f"""
        <html><body style="font-family:Arial;padding:20px;background:#f5f5f5;">
        <div style="max-width:600px;margin:0 auto;background:#fff;padding:24px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
            <h2 style="color:#f44336;margin:0 0 16px;border-bottom:2px solid #f44336;padding-bottom:8px;">Absence Notification</h2>
            <p>Hi <strong>{user_name}</strong>,</p>
            <p>No morning check-in found for <strong>{dt.date.today().strftime("%B %d, %Y")}</strong>.</p>
            <p>If you're on approved leave or this is an error, please contact HR immediately.</p>
            <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
            <p style="color:#666;font-size:12px;">Automated message • {Config.APP_TITLE}</p>
        </div></body></html>
        """
    return subject, body

def ensure_user_columns():
    sess = SessionLocal()
    try:
        for col in ["email", "external_id", "role"]:
            try:
                sess.execute(text(f"SELECT {col} FROM users LIMIT 1"))
            except Exception:
                try:
                    col_type = "VARCHAR(255)" if col != "role" else "VARCHAR(20) DEFAULT 'user'"
                    sess.execute(text(f"ALTER TABLE users ADD COLUMN {col} {col_type}"))
                    sess.commit()
                    logger.info(f"Added column '{col}' to users table")
                except Exception as e:
                    logger.warning(f"Could not add column '{col}': {e}")
    finally:
        sess.close()

def compute_metrics(sess):
    total_users = sess.query(User).count()
    today = today_str()
    recs = sess.query(Attendance).filter(Attendance.date == today).all()
    present = len([r for r in recs if r.checkin1_time])
    late = len([r for r in recs if r.checkin1_time and is_late(1, r.checkin1_time)])
    on_time = present - late
    absent = total_users - present
    return total_users, present, absent, late, on_time

# ============================================================================
# STATIC FILES
# ============================================================================

@app.get("/seed/<path:filename>")
def seed_image(filename):
    safe = os.path.normpath(filename).replace("\\", "/")
    if safe.startswith("../") or safe.startswith("/"):
        return abort(400)
    return send_from_directory(Config.SEED_DIR, safe, as_attachment=False)

@app.get("/last_frame.jpg")
def last_frame():
    if not os.path.exists(Config.ENGINE_FRAME_PATH):
        return ("", 404)
    resp = make_response(send_file(Config.ENGINE_FRAME_PATH, mimetype="image/jpeg"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# ============================================================================
# AUTH
# ============================================================================

@app.get("/login")
def login():
    if session.get("logged_in"):
        return redirect(url_for("home"))
    return render_template("login.html", title=Config.APP_TITLE)

@app.post("/login")
@limiter.limit("5 per minute")
def do_login():
    password = (request.form.get("password") or "").strip()
    
    if check_password_hash(Config.ADMIN_PASSWORD_HASH, password):
        session.permanent = True
        session["logged_in"] = True
        session["is_admin"] = True
        session["username"] = "admin"
        logger.info(f"Successful login from {get_remote_address()}")
        
        next_url = request.args.get("next")
        if next_url and next_url.startswith("/"):
            return redirect(next_url)
        return redirect(url_for("home"))
    
    logger.warning(f"Failed login attempt from {get_remote_address()}")
    flash("Invalid password", "error")
    return redirect(url_for("login"))

@app.get("/logout")
@require_auth
def logout():
    logger.info(f"User {session.get('username')} logged out")
    session.clear()
    flash("You have been logged out", "success")
    return redirect(url_for("login"))

# ============================================================================
# CONTEXT PROCESSOR
# ============================================================================

@app.context_processor
def inject_globals():
    return {
        "APP_TITLE": Config.APP_TITLE,
        "now": dt.datetime.now,
        "gates": GATES,
        "logged_in": session.get("logged_in", False),
        "is_admin": session.get("is_admin", False),
        "username": session.get("username", "Guest")
    }

# ============================================================================
# DASHBOARD
# ============================================================================

@app.get("/")
@require_auth
def home():
    sess = SessionLocal()
    try:
        total_users, present, absent, late, on_time = compute_metrics(sess)
        
        week_ago = (dt.date.today() - dt.timedelta(days=7)).isoformat()
        recent = sess.query(Attendance).filter(Attendance.date >= week_ago).all()
        
        daily_stats = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0})
        for r in recent:
            date = r.date
            if r.checkin1_time:
                daily_stats[date]["present"] += 1
                if is_late(1, r.checkin1_time):
                    daily_stats[date]["late"] += 1
        
        today = today_str()
        today_records = sess.query(Attendance, User).join(
            User, Attendance.user_id == User.user_id
        ).filter(Attendance.date == today).all()
        
        late_employees = [
            {"name": u.name, "time": a.checkin1_time}
            for a, u in today_records
            if a.checkin1_time and is_late(1, a.checkin1_time)
        ]
        
        all_users = sess.query(User).all()
        present_ids = {a.user_id for a, _ in today_records if a.checkin1_time}
        absent_employees = [
            {"name": u.name, "id": u.user_id}
            for u in all_users
            if u.user_id not in present_ids
        ]
        
        current_quarter = quarter_for(dt.date.today())
        quarter_records = sess.query(Attendance).filter(
            Attendance.quarter_id == current_quarter
        ).all()
        
        total_checkins = len(quarter_records) * 4
        completed_checkins = sum([
            1 for r in quarter_records if r.checkin1_time
        ] + [
            1 for r in quarter_records if r.checkin2_time
        ] + [
            1 for r in quarter_records if r.checkin3_time
        ] + [
            1 for r in quarter_records if r.checkin4_time
        ])
        
        quarterly_compliance = round((completed_checkins / total_checkins * 100) if total_checkins > 0 else 0, 1)
        
        metrics = {
            "total_users": total_users,
            "present_today": present,
            "absent_today": absent,
            "late_today": late,
            "on_time_today": on_time,
            "attendance_rate": round((present / total_users * 100) if total_users > 0 else 0, 1),
            "late_employees": late_employees[:5],
            "absent_employees": absent_employees[:5],
            "weekly_labels": sorted(daily_stats.keys())[-7:],
            "weekly_present": [daily_stats[d]["present"] for d in sorted(daily_stats.keys())[-7:]],
            "weekly_late": [daily_stats[d]["late"] for d in sorted(daily_stats.keys())[-7:]],
            "quarterly_compliance": quarterly_compliance,
            "current_quarter": int(current_quarter[-1]),
            "current_year": int(current_quarter[:4])
        }
        
    finally:
        sess.close()
    
    return render_template(
        "dashboard_enhanced.html",
        title=Config.APP_TITLE,
        today=dt.date.today(),
        metrics=metrics,
        all_users=all_users,
        gates=GATES,
        enable_stream=Config.ENABLE_CAMERA_STREAM,
    )

# ============================================================================
# LIVE CAMERA API
# ============================================================================

@app.get("/video_feed")
def video_feed():
    def generate():
        while True:
            if os.path.exists(Config.ENGINE_FRAME_PATH):
                try:
                    with open(Config.ENGINE_FRAME_PATH, 'rb') as f:
                        frame = f.read()
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n'
                           b'Cache-Control: no-store\r\n\r\n' + frame + b'\r\n')
                except Exception:
                    pass
            time.sleep(0.1)
    
    resp = Response(generate(), mimetype='multipart/x-mixed-replace; boundary=frame')
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.get("/api/engine_status")
def api_engine_status():
    if not os.path.exists(Config.ENGINE_STATUS_PATH):
        return jsonify(ok=False, message="Engine status not available")
    
    try:
        with open(Config.ENGINE_STATUS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(ok=True, **data)
    except Exception as e:
        logger.error(f"Error reading engine status: {e}")
        return jsonify(ok=False, error=str(e))

# ============================================================================
# INDIVIDUAL ANALYTICS API
# ============================================================================

@app.get("/api/user_analytics/<int:user_id>")
@require_auth
def api_user_analytics(user_id):
    sess = SessionLocal()
    try:
        user = sess.get(User, user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        end_date = dt.date.today()
        start_date = end_date - dt.timedelta(days=30)
        
        records = sess.query(Attendance).filter(
            Attendance.user_id == user_id,
            Attendance.date >= start_date.isoformat(),
            Attendance.date <= end_date.isoformat()
        ).all()
        
        total_days = 30
        missing_g1 = sum(1 for r in records if not r.checkin1_time)
        missing_g2 = sum(1 for r in records if not r.checkin2_time)
        missing_g3 = sum(1 for r in records if not r.checkin3_time)
        missing_g4 = sum(1 for r in records if not r.checkin4_time)
        
        total_possible = total_days * 4
        total_completed = sum([
            1 for r in records if r.checkin1_time
        ] + [
            1 for r in records if r.checkin2_time
        ] + [
            1 for r in records if r.checkin3_time
        ] + [
            1 for r in records if r.checkin4_time
        ])
        
        compliance_rate = round((total_completed / total_possible * 100) if total_possible > 0 else 0, 1)
        
        heatmap = []
        for i in range(30):
            day_date = end_date - dt.timedelta(days=29 - i)
            rec = next((r for r in records if r.date == day_date.isoformat()), None)
            
            if rec and rec.checkin1_time:
                status = 'late' if is_late(1, rec.checkin1_time) else 'present'
            else:
                status = 'absent'
            
            heatmap.append({
                "day": day_date.day,
                "date": day_date.isoformat(),
                "status": status
            })
        
        return jsonify({
            "compliance_rate": compliance_rate,
            "missing_g1": missing_g1,
            "missing_g2": missing_g2,
            "missing_g3": missing_g3,
            "missing_g4": missing_g4,
            "heatmap": heatmap
        })
        
    finally:
        sess.close()

# ============================================================================
# USER MANAGEMENT
# ============================================================================

@app.get("/users")
@require_auth
def users_page():
    sess = SessionLocal()
    users = sess.query(User).order_by(User.name).all()
    sess.close()
    return render_template("users_manage.html", title=Config.APP_TITLE, users=users)

@app.post("/users/add")
@require_auth
@admin_required
@limiter.limit("10 per hour")
def users_add():
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip() or None
    extid = (request.form.get("external_id") or "").strip() or None
    file = request.files.get("photo")
    
    if not name or not file or not file.filename:
        flash("Name and photo are required", "error")
        return redirect(url_for("users_page"))
    
    sess = SessionLocal()
    dst = None
    
    try:
        ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
        dst = os.path.join(Config.SEED_DIR, f"{name}{ext}")
        file.save(dst)
        logger.info(f"Saved photo to: {dst}")
        
        import numpy as np
        import cv2
        from PIL import Image, ImageOps
        from deepface import DeepFace
        
        def read_bgr(pth):
            im = Image.open(pth)
            im = ImageOps.exif_transpose(im)
            arr = np.array(im.convert("RGB"))
            return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        
        model_name = os.getenv("MODEL_NAME", "Facenet")
        logger.info(f"Processing face recognition for {name}")
        
        rep = DeepFace.represent(
            img_path=read_bgr(dst),
            model_name=model_name,
            detector_backend=os.getenv("DETECTOR", "opencv"),
            enforce_detection=True,
            align=True
        )
        emb = rep[0]["embedding"]
        logger.info(f"Face embedding generated successfully for {name}")
        
        u = sess.query(User).filter_by(name=name).one_or_none()
        if u:
            u.face_embedding = json.dumps(emb)
            if hasattr(u, "email"):
                u.email = email
            if hasattr(u, "external_id"):
                u.external_id = extid
            logger.info(f"Updated user: {name}")
            flash(f"Successfully updated {name}", "success")
        else:
            kwargs = dict(name=name, face_embedding=json.dumps(emb))
            if "email" in User.__table__.columns.keys():
                kwargs["email"] = email
            if "external_id" in User.__table__.columns.keys():
                kwargs["external_id"] = extid
            u = User(**kwargs)
            sess.add(u)
            logger.info(f"Added new user: {name}")
            flash(f"Successfully enrolled {name}", "success")
        
        sess.commit()
        
        if email and Config.ENABLE_EMAIL and not u.user_id:
            threading.Thread(
                target=send_welcome_email,
                args=(name, email),
                daemon=True
            ).start()
        
    except Exception as e:
        sess.rollback()
        logger.error(f"Enrollment failed for {name}: {e}", exc_info=True)
        
        if dst and os.path.exists(dst):
            try:
                os.remove(dst)
                logger.info(f"Removed failed upload: {dst}")
            except Exception as cleanup_error:
                logger.error(f"Failed to cleanup image: {cleanup_error}")
        
        error_msg = str(e)
        if "No face" in error_msg or "Face could not be detected" in error_msg:
            flash("No face detected in the photo. Please upload a clear photo with a visible face.", "error")
        elif "Multiple faces" in error_msg or "more than one face" in error_msg:
            flash("Multiple faces detected. Please upload a photo with only one person.", "error")
        else:
            flash(f"Enrollment failed: {error_msg}", "error")
    
    finally:
        sess.close()
    
    return redirect(url_for("users_page"))

@app.post("/users/update/<int:user_id>")
@require_auth
@admin_required
def user_update(user_id):
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip() or None
    extid = (request.form.get("external_id") or "").strip() or None
    
    sess = SessionLocal()
    u = sess.get(User, user_id)
    if u:
        if name:
            u.name = name
        if hasattr(u, "email"):
            u.email = email
        if hasattr(u, "external_id"):
            u.external_id = extid
        sess.commit()
        logger.info(f"Updated user ID {user_id}: {name}")
        flash("User updated successfully", "success")
    else:
        flash("User not found", "error")
    sess.close()
    return redirect(url_for("users_page"))

@app.post("/users/delete/<int:user_id>")
@require_auth
@admin_required
def users_delete(user_id):
    sess = SessionLocal()
    u = sess.get(User, user_id)
    if u:
        sess.delete(u)
        sess.commit()
        logger.info(f"Deleted user: {u.name} (ID: {user_id})")
        flash("User deleted successfully", "success")
    else:
        flash("User not found", "error")
    sess.close()
    return redirect(url_for("users_page"))

# ============================================================================
# ATTENDANCE
# ============================================================================

@app.get("/attendance")
@require_auth
def attendance_page():
    d = request.args.get("date") or today_str()
    sess = SessionLocal()
    
    rows = (
        sess.query(Attendance, User)
        .join(User, Attendance.user_id == User.user_id)
        .filter(Attendance.date == d)
        .order_by(User.name)
        .all()
    )
    
    def seed_ext_for(name: str) -> Optional[str]:
        for ext in (".jpg", ".jpeg", ".png", ".JPG", ".PNG"):
            p = os.path.join(Config.SEED_DIR, f"{name}{ext}")
            if os.path.exists(p):
                return ext
        return None
    
    data = []
    for att, user in rows:
        def badge(val, gid):
            if not val:
                return {"text": "-", "cls": "muted"}
            return {"text": val, "cls": ("late" if is_late(gid, val) else "ok")}
        
        ext = seed_ext_for(user.name)
        photo_url = url_for('seed_image', filename=f"{user.name}{ext}") if ext else None
        data.append({
            "id": user.user_id,
            "name": user.name,
            "email": getattr(user, "email", None),
            "photo": photo_url,
            "g1": badge(att.checkin1_time, 1),
            "g2": badge(att.checkin2_time, 2),
            "g3": badge(att.checkin3_time, 3),
            "g4": badge(att.checkin4_time, 4),
            "date": att.date
        })
    
    users = sess.query(User).order_by(User.name).all()
    sess.close()
    
    return render_template(
        "attendance_sheet.html",
        title=Config.APP_TITLE,
        date=d,
        rows=data,
        users=users,
        gates=GATES
    )

@app.post("/attendance/edit")
@require_auth
@admin_required
def attendance_edit():
    user_id = int(request.form["user_id"])
    date = request.form.get("date") or today_str()
    t1 = request.form.get("g1") or None
    t2 = request.form.get("g2") or None
    t3 = request.form.get("g3") or None
    t4 = request.form.get("g4") or None
    
    sess = SessionLocal()
    att = sess.query(Attendance).filter(
        and_(Attendance.user_id == user_id, Attendance.date == date)
    ).one_or_none()
    
    if not att:
        att = Attendance(
            user_id=user_id,
            date=date,
            quarter_id=quarter_for(dt.date.fromisoformat(date))
        )
        sess.add(att)
        sess.flush()
    
    att.checkin1_time = t1
    att.checkin2_time = t2
    att.checkin3_time = t3
    att.checkin4_time = t4
    sess.commit()
    sess.close()
    
    logger.info(f"Attendance edited for user {user_id} on {date}")
    flash("Attendance updated successfully", "success")
    return redirect(url_for("attendance_page", date=date))
#===========================================================================
#charts
#===========================================================================
@app.get("/charts")
@require_auth
def charts_page():
    sess = SessionLocal()
    try:
        end_date = dt.date.today()
        start_date = end_date - dt.timedelta(days=30)
        
        records = sess.query(Attendance, User).join(
            User, Attendance.user_id == User.user_id
        ).filter(
            Attendance.date >= start_date.isoformat(),
            Attendance.date <= end_date.isoformat()
        ).all()
        
        total_users = sess.query(User).count()
        today = today_str()
        today_records = sess.query(Attendance).filter(Attendance.date == today).all()
        
        present_count = len([r for r in today_records if r.checkin1_time])
        late_count = len([r for r in today_records if r.checkin1_time and is_late(1, r.checkin1_time)])
        on_time_count = present_count - late_count
        absent_count = total_users - present_count
        
        pie_data = {
            "labels": ["On Time", "Late", "Absent"],
            "values": [int(on_time_count), int(late_count), int(absent_count)],
            "colors": ["#10b981", "#f59e0b", "#ef4444"]
        }
        
        bar_labels = []
        bar_present = []
        bar_late = []
        
        for i in range(7):
            day = end_date - dt.timedelta(days=6-i)
            bar_labels.append(day.strftime("%m/%d"))
            
            day_records = [a for a, u in records if a.date == day.isoformat()]
            present = len([r for r in day_records if r.checkin1_time])
            late = len([r for r in day_records if r.checkin1_time and is_late(1, r.checkin1_time)])
            
            bar_present.append(int(present))
            bar_late.append(int(late))
        
        bar_data = {
            "labels": bar_labels,
            "present": bar_present,
            "late": bar_late
        }
        
        heatmap_data = []
        for i in range(30):
            day = end_date - dt.timedelta(days=29-i)
            day_records = [a for a, u in records if a.date == day.isoformat()]
            
            if day_records:
                total_checkins = len([1 for a in day_records if a.checkin1_time])
                rate = (total_checkins / total_users * 100) if total_users > 0 else 0
            else:
                rate = 0
            
            heatmap_data.append({
                "date": day.isoformat(),
                "day": day.day,
                "month": day.strftime("%b"),
                "rate": round(rate, 1),
                "weekday": day.strftime("%a")
            })
        
    finally:
        sess.close()
    
    return render_template(
        "charts.html",
        title=Config.APP_TITLE,
        pie_data=pie_data,
        bar_data=bar_data,
        heatmap_data=heatmap_data,
        today=end_date
    )

# ============================================================================
# EMPLOYEES
# ============================================================================

@app.get("/employees")
@require_auth
def employees_search():
    search_name = request.args.get("search", "").strip()
    period = request.args.get("period", "week")
    
    if period == "week":
        start = dt.date.today() - dt.timedelta(days=dt.date.today().weekday())
        end = start + dt.timedelta(days=6)
    elif period == "month":
        start = dt.date.today().replace(day=1)
        nm = (start.month % 12) + 1
        year = start.year + (1 if nm == 1 else 0)
        end = dt.date(year, nm, 1) - dt.timedelta(days=1)
    else:
        start = dt.date(2020, 1, 1)
        end = dt.date.today()
    
    sess = SessionLocal()
    query = sess.query(User).order_by(User.name)
    if search_name:
        query = query.filter(User.name.ilike(f"%{search_name}%"))
    users = query.all()
    
    results = []
    for user in users:
        attendance_records = sess.query(Attendance).filter(
            Attendance.user_id == user.user_id,
            Attendance.date >= start.isoformat(),
            Attendance.date <= end.isoformat()
        ).all()
        
        total_days = (end - start).days + 1
        present = len([r for r in attendance_records if r.checkin1_time])
        absent = total_days - present
        late = len([r for r in attendance_records if r.checkin1_time and is_late(1, r.checkin1_time)])
        
        results.append({
            "id": user.user_id,
            "name": user.name,
            "email": getattr(user, "email", ""),
            "external_id": getattr(user, "external_id", ""),
            "present": present,
            "absent": absent,
            "late": late,
            "total_days": total_days,
            "attendance_rate": round((present / total_days * 100) if total_days > 0 else 0, 1)
        })
    
    sess.close()
    return render_template(
        "employees.html",
        title=Config.APP_TITLE,
        employees=results,
        search=search_name,
        period=period,
        start=start,
        end=end
    )

# ============================================================================
# EXPORT
# ============================================================================

@app.get("/export/daily-excel")
@require_auth
@limiter.limit("10 per hour")
def export_daily():
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    
    d = request.args.get("date") or today_str()
    sess = SessionLocal()
    
    rows = (
        sess.query(Attendance, User)
        .join(User, Attendance.user_id == User.user_id)
        .filter(Attendance.date == d)
        .order_by(User.name)
        .all()
    )
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {d}"
    
    headers = ["Photo", "Name", "Email", "ID", "Check-in 1", "Check-in 2", "Check-in 3", "Check-in 4", "Status", "Scanned"]
    ws.append(headers)
    
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    
    def find_photo(name: str):
        for ext in (".jpg", ".jpeg", ".png", ".JPG", ".PNG"):
            p = os.path.join(Config.SEED_DIR, f"{name}{ext}")
            if os.path.exists(p):
                return p
        return None
    
    r_idx = 2
    for att, user in rows:
        ws.row_dimensions[r_idx].height = 64
        
        photo = find_photo(user.name)
        if photo:
            try:
                img = XLImage(photo)
                img.height = 48
                img.width = 48
                ws.add_image(img, f"A{r_idx}")
            except Exception:
                pass
        
        ws.cell(r_idx, 2, user.name)
        ws.cell(r_idx, 3, getattr(user, "email", "") or "")
        ws.cell(r_idx, 4, getattr(user, "external_id", "") or "")
        
        g1, g2, g3, g4 = att.checkin1_time, att.checkin2_time, att.checkin3_time, att.checkin4_time
        ws.cell(r_idx, 5, g1 or "-")
        ws.cell(r_idx, 6, g2 or "-")
        ws.cell(r_idx, 7, g3 or "-")
        ws.cell(r_idx, 8, g4 or "-")
        
        status = "ABSENT" if not g1 else ("LATE" if is_late(1, g1) else "On-time")
        ws.cell(r_idx, 9, status)
        ws.cell(r_idx, 10, "Yes" if g1 else "No")
        r_idx += 1
    
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    b = io.BytesIO()
    wb.save(b)
    b.seek(0)
    sess.close()
    
    logger.info(f"Exported attendance for {d}")
    return send_file(
        b,
        as_attachment=True,
        download_name=f"attendance_{d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================================
# HEALTH & ERRORS
# ============================================================================

@app.get("/health")
def health_check():
    try:
        sess = SessionLocal()
        sess.execute(text("SELECT 1"))
        sess.close()
        db_status = "ok"
    except Exception as e:
        logger.error(f"Health check DB error: {e}")
        db_status = "error"
    
    engine_status = "ok" if os.path.exists(Config.ENGINE_FRAME_PATH) else "offline"
    
    return jsonify({
        "status": "healthy" if db_status == "ok" else "degraded",
        "timestamp": dt.datetime.now().isoformat(),
        "database": db_status,
        "engine": engine_status,
        "version": "1.0.0"
    })

@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", error_code=404, error_message="Page not found"), 404

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal error: {e}")
    return render_template("error.html", error_code=500, error_message="Internal server error"), 500

@app.errorhandler(429)
def ratelimit_handler(e):
    return jsonify(error="Rate limit exceeded. Please try again later."), 429

# ============================================================================
# BACKGROUND WATCHER
# ============================================================================

class Watcher(threading.Thread):
    def __init__(self):
        super().__init__(daemon=True)
        self.seen = {}
        self.absent_alert_sent = set()
        self._day = today_str()

    def run(self):
        logger.info("Watcher thread started")
        while True:
            try:
                if today_str() != self._day:
                    self._day = today_str()
                    self.absent_alert_sent.clear()
                    self.seen.clear()
                    logger.info("Daily reset completed")

                sess = SessionLocal()
                rows = sess.query(Attendance).filter(Attendance.date == self._day).all()
                now = dt.datetime.now().time()

                for r in rows:
                    key = (r.user_id, r.date)
                    cur = (r.checkin1_time, r.checkin2_time, r.checkin3_time, r.checkin4_time)
                    prev = self.seen.get(key)
                    
                    if prev and prev != cur:
                        user = sess.get(User, r.user_id)
                        if cur[0] and is_late(1, cur[0]) and (not prev[0] or not is_late(1, prev[0])):
                            subj, body = email_template("late", user.name, cur[0])
                            send_email(getattr(user, "email", None), subj, body)
                    
                    self.seen[key] = cur

                if now >= dt.time(10, 0):
                    all_users = sess.query(User).all()
                    present_ids = {r.user_id for r in rows if r.checkin1_time}
                    
                    for u in all_users:
                        if u.user_id in present_ids or u.user_id in self.absent_alert_sent:
                            continue
                        
                        subj, body = email_template("absent", u.name)
                        send_email(getattr(u, "email", None), subj, body)
                        self.absent_alert_sent.add(u.user_id)

                sess.close()
            except Exception as e:
                logger.error(f"Watcher error: {e}")
            
            time.sleep(30)

watcher = None

def start_watch():
    global watcher
    if watcher is None and Config.ENABLE_EMAIL:
        watcher = Watcher()
        watcher.start()
        logger.info("Email watcher enabled")
    else:
        logger.info("Email watcher disabled (SMTP not configured)")

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    init_db()
    ensure_user_columns()
    start_watch()
    
    logger.info(f"Starting {Config.APP_TITLE} in {Config.ENV} mode")
    logger.info(f"Debug mode: {Config.DEBUG}")
    
    if Config.ENV == "production":
        logger.warning("Running with Flask dev server in production mode. Use Gunicorn/uWSGI instead!")
    
    app.run(
        debug=Config.DEBUG,
        threaded=True,
        host="0.0.0.0",
        port=int(os.getenv("PORT", 5000))
    )